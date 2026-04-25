"""
build_sheet_processor.py

Reads a MASTER CUTSHEET Excel file (e.g., MASTER-US-CENTRAL-08A-...-ELLENDALE)
and a MASTER REGION TEMPLATE, then returns all cable/device/optic data for a
given room designator + rack number.

Mirrors the logic of Build Sheet V2.2.1.xlsx.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

from utils import natural_key as _natural_key

# ---------------------------------------------------------------------------
# Column mappings  (Excel header → internal field name)
# ---------------------------------------------------------------------------

CUTSHEET_COLS = {
    'STATUS':                      'status',
    'A-SIDE LOCODE':               'a_locode',
    'A-LOC:CAB:RU':                'a_loc',
    'A-SIDE-DNS-NAME':             'a_dns',
    'A-MODEL':                     'a_model',
    'A-PORT':                      'a_port',
    'A-BREAKOUT LOC:CAB:RU':       'a_breakout_loc',
    'A-BREAKOUT SLOT:PORT':        'a_breakout_port',
    'A-OPTIC':                     'a_optic',
    'A-PATCH-PANEL LOC:CAB:RU:PORT': 'a_patch',
    'Z-SIDE LOCODE':               'z_locode',
    'Z-LOC:CAB:RU':                'z_loc',
    'Z-SIDE-DNS-NAME':             'z_dns',
    'Z-MODEL':                     'z_model',
    'Z-PORT':                      'z_port',
    'Z-BREAKOUT LOC:CAB:RU':       'z_breakout_loc',
    'Z-BREAKOUT SLOT:PORT':        'z_breakout_port',
    'Z-OPTIC':                     'z_optic',
    'Z-PATCH-PANEL LOC:CAB:RU:PORT': 'z_patch',
    'CABLE':                       'cable_type',
}

HOSTS_COLS = {
    'STATUS':           'status',
    'LAST-PROVISIONED': 'last_provisioned',
    'LOC:CAB:RU':       'loc',
    'DNS-A-RECORD':     'dns',
    'NETBOX MODEL':     'model',
    'SERIAL':           'serial',
    'ROLE':             'role',
    'ROW:TYPE':         'row_type',
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _str(val):
    return str(val).strip() if val is not None else ''


def _read_sheet(wb, sheet_name, col_map):
    """
    Read a worksheet into a list of dicts, mapping headers via col_map.
    Returns [] if the sheet does not exist.
    """
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return []

    headers = [_str(h).upper() for h in header_row]
    col_indices = {}
    for col_name, field in col_map.items():
        key = col_name.upper()
        if key in headers:
            col_indices[field] = headers.index(key)

    result = []
    for row in rows:
        if all(v is None for v in row):
            continue
        record = {field: _str(row[idx]) if idx < len(row) else ''
                  for field, idx in col_indices.items()}
        result.append(record)
    return result


def _find_cutsheet_tab(wb):
    """Find the cutsheet tab name, case-insensitive. Falls back to column heuristic."""
    for name in wb.sheetnames:
        if name.strip().lower() in ('cutsheet', 'connections'):
            return name
    # Fallback: look for a sheet with A-OPTIC / Z-OPTIC columns
    required = {'A-OPTIC', 'Z-OPTIC'}
    for name in wb.sheetnames:
        ws = wb[name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header:
            cols = {_str(h).upper() for h in header if h is not None}
            if required.issubset(cols):
                return name
    return None


def _parse_loc(loc_str):
    """
    Parse a location string like 'dh202:043:35' into (room, rack, ru).
    Returns ('', '', '') on failure.
    """
    if not loc_str:
        return '', '', ''
    parts = loc_str.lower().split(':')
    room = parts[0] if len(parts) > 0 else ''
    rack = parts[1].lstrip('0') or '0' if len(parts) > 1 else ''
    ru   = parts[2] if len(parts) > 2 else ''
    return room, rack, ru


def _room_matches(room_code, room_filter):
    """
    room_filter 'dh2' matches 'dh202', 'dh201', etc.
    room_filter 'dh202' matches exactly 'dh202'.
    """
    return room_code == room_filter or room_code.startswith(room_filter)


def _rack_matches(rack_code, rack_filter):
    """Compare rack codes after stripping leading zeros."""
    return (rack_code.lstrip('0') or '0') == (rack_filter.lstrip('0') or '0')


def _loc_in_rack(loc_str, room_filter, rack_filter):
    room, rack, _ = _parse_loc(loc_str)
    if not room:
        return False
    return _room_matches(room, room_filter) and _rack_matches(rack, rack_filter)


def _loc_in_room(loc_str, room_filter):
    room, _, _ = _parse_loc(loc_str)
    if not room:
        return False
    return _room_matches(room, room_filter)


def _resolve_room_filter(cables_raw, hosts_by_loc, room_filter, rack_filter):
    """
    Resolve shorthand room filters like 'dh2' to an exact hall when possible.
    If more than one exact hall matches the requested rack, raise a clear error
    instead of silently combining multiple halls into one rack view.
    """
    if not room_filter or len(room_filter) >= 5:
        return room_filter

    matches = set()

    for cable in cables_raw:
        for loc_key in ('a_loc', 'z_loc'):
            room_code, rack_code, _ = _parse_loc(cable.get(loc_key, ''))
            if not room_code or not _rack_matches(rack_code, rack_filter):
                continue
            if _room_matches(room_code, room_filter):
                matches.add(room_code)

    for loc in hosts_by_loc:
        room_code, rack_code, _ = _parse_loc(loc)
        if not room_code or not _rack_matches(rack_code, rack_filter):
            continue
        if _room_matches(room_code, room_filter):
            matches.add(room_code)

    if len(matches) <= 1:
        return next(iter(matches), room_filter)

    raise ValueError(
        f"Ambiguous room '{room_filter}' for rack {rack_filter}. "
        f"Matches: {', '.join(sorted(matches))}. Enter an exact hall."
    )


def _derive_device_status(statuses):
    """Collapse cable statuses into a device-level status label."""
    cleaned = [s.strip() for s in statuses if s and str(s).strip()]
    if not cleaned:
        return ''
    lowered = [s.lower() for s in cleaned]
    if all(s.startswith('cable not run') for s in lowered):
        return 'Pending'
    return 'Installed'


def _ru_sort(ru_str):
    try:
        return int(ru_str)
    except (ValueError, TypeError):
        return 0


def _cable_label(cable, perspective_side):
    """
    Generate a human-readable label string for a cable.
    perspective_side = 'a' means this rack is on the A side.
    """
    if perspective_side == 'a':
        src = f"{cable.get('a_loc','')} port {cable.get('a_port','')}"
        dst = f"{cable.get('z_loc','')} port {cable.get('z_port','')}"
    else:
        src = f"{cable.get('z_loc','')} port {cable.get('z_port','')}"
        dst = f"{cable.get('a_loc','')} port {cable.get('a_port','')}"
    cable_type = cable.get('cable_type', '')
    parts = [f"{src}  →  {dst}"]
    if cable_type:
        parts.append(f"Cable: {cable_type}")
    return '  |  '.join(parts)


# ---------------------------------------------------------------------------
# Cab type + elevation helpers
# ---------------------------------------------------------------------------

def _overhead_rack_to_cab_type(cutsheet_path):
    """
    Load the OVERHEAD sheet once and return a dict of {rack_int: cab_type_str}
    for every rack that has a valid (non-RES) cab type directly below it.
    """
    wb = openpyxl.load_workbook(cutsheet_path, read_only=True, data_only=True)
    if 'OVERHEAD' not in wb.sheetnames:
        wb.close()
        return {}
    data = [row for row in wb['OVERHEAD'].iter_rows(values_only=True)]
    wb.close()

    mapping = {}
    for r, row in enumerate(data):
        for c, val in enumerate(row):
            if isinstance(val, (int, float)) and val == int(val) and r + 1 < len(data):
                below = data[r + 1][c]
                if isinstance(below, str) and below.strip() and below.strip().upper() != 'RES':
                    rack_int = int(val)
                    if rack_int not in mapping:
                        mapping[rack_int] = below.strip()
    return mapping


def _lookup_cab_type(cutsheet_path, rack_filter):
    """Return the cab type for a single rack number."""
    try:
        rack_num = int(rack_filter.lstrip('0') or '0')
    except (ValueError, TypeError):
        return ''
    return _overhead_rack_to_cab_type(cutsheet_path).get(rack_num, '')


def _cab_type_summary(cutsheet_path, cables_raw, room_filter):
    """
    Return {cab_type: rack_count} for all racks in the given room that
    have at least one cable entry in the cutsheet. Only cab types with
    count > 0 are included.
    """
    rack_to_cab = _overhead_rack_to_cab_type(cutsheet_path)

    # Collect unique rack numbers present in cable data for this room
    unique_racks = set()
    for cable in cables_raw:
        for loc_key in ('a_loc', 'z_loc'):
            room, rack, _ = _parse_loc(cable.get(loc_key, ''))
            if room and _room_matches(room, room_filter) and rack:
                try:
                    unique_racks.add(int(rack.lstrip('0') or '0'))
                except ValueError:
                    pass

    counts = defaultdict(int)
    for rack_int in unique_racks:
        cab = rack_to_cab.get(rack_int)
        if cab:
            counts[cab] += 1

    return dict(sorted(counts.items(), key=lambda x: x[1], reverse=True))


def _lookup_elevation(template_path, cab_type):
    """
    Open the MASTER-REGION-TEMPLATE and return the elevation rows for
    the given cab type from the sheet named 'ELEV {cab_type}'.
    """
    sheet_name = f'ELEV {cab_type}'
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    data = [row for row in wb[sheet_name].iter_rows(values_only=True)]
    wb.close()

    # Find header row (contains RU-BASE and DEVICE NAME)
    header_idx = next(
        (i for i, row in enumerate(data)
         if any(_str(v).upper() in ('RU-BASE', 'DEVICE NAME') for v in row if v is not None)),
        None
    )
    if header_idx is None:
        return []

    headers = [_str(v).upper() for v in data[header_idx]]
    col = {h: i for i, h in enumerate(headers)}
    ru_col   = col.get('RU-BASE')
    name_col = col.get('DEVICE NAME')
    type_col = col.get('NETBOX-DEVICE-TYPE')

    elevation = []
    for row in data[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        ru_raw = row[ru_col] if ru_col is not None and ru_col < len(row) else None
        # Strip .0 from float RU values (Excel stores integers as floats)
        if isinstance(ru_raw, float) and ru_raw == int(ru_raw):
            ru = str(int(ru_raw))
        else:
            ru = _str(ru_raw)
        name = _str(row[name_col]) if name_col is not None and name_col < len(row) else ''
        dev_type = _str(row[type_col]) if type_col is not None and type_col < len(row) else ''
        if ru or name:
            elevation.append({'ru': ru, 'device_name': name, 'device_type': dev_type})

    return elevation


# ---------------------------------------------------------------------------
# Main processor
# ---------------------------------------------------------------------------

def process_rack(cutsheet_path, template_path, room_input, rack_input):
    """
    Load the CUTSHEET (and optionally SITE-HOSTS) from cutsheet_path,
    filter to the given room + rack, and return structured results.

    Returns a dict ready to be serialised to JSON.
    """
    room_filter = room_input.strip().lower()
    rack_filter  = rack_input.strip()

    # -- Load cutsheet workbook --
    wb_cut = openpyxl.load_workbook(cutsheet_path, read_only=True, data_only=True)
    cut_tab = _find_cutsheet_tab(wb_cut) or 'CUTSHEET'
    cables_raw = _read_sheet(wb_cut, cut_tab, CUTSHEET_COLS)

    # -- Enrich from SITE-HOSTS if present --
    hosts_by_loc = {}
    hosts_raw = _read_sheet(wb_cut, 'SITE-HOSTS', HOSTS_COLS)
    for h in hosts_raw:
        loc = h.get('loc', '')
        if loc:
            hosts_by_loc[loc] = h
    wb_cut.close()

    room_filter = _resolve_room_filter(cables_raw, hosts_by_loc, room_filter, rack_filter)

    # -- Cab type from OVERHEAD + elevation from template --
    cab_type      = _lookup_cab_type(cutsheet_path, rack_filter)
    elevation     = _lookup_elevation(template_path, cab_type) if cab_type and template_path else []
    cab_type_summary = _cab_type_summary(cutsheet_path, cables_raw, room_filter)

    # -- Filter cables where this rack is the A-side only --
    # Excluding Z-side-only cables prevents duplicate links when two racks
    # generate cable maps: each link is owned by its A-side rack.
    rack_cables = []
    for cable in cables_raw:
        a_in = _loc_in_rack(cable.get('a_loc', ''), room_filter, rack_filter)
        if not a_in:
            continue
        z_in = _loc_in_rack(cable.get('z_loc', ''), room_filter, rack_filter)
        cable['a_in_rack'] = True
        cable['z_in_rack'] = z_in
        rack_cables.append(cable)

    # -- Split internal vs cab-to-cab --
    internal     = [c for c in rack_cables if c['a_in_rack'] and c['z_in_rack']]
    cab_to_cab   = [c for c in rack_cables if not (c['a_in_rack'] and c['z_in_rack'])]

    # -- Build per-location cable status list from both sides --
    loc_statuses = defaultdict(list)
    devices = {}

    def _upsert_device(loc, ru, dns_name='', model='', status=''):
        if not loc:
            return
        current = devices.get(loc)
        if current is None:
            devices[loc] = {
                'location': loc,
                'ru': ru,
                'dns_name': dns_name or '',
                'model': model or '',
                'status': status or '',
            }
            return
        if not current.get('ru') and ru:
            current['ru'] = ru
        if not current.get('dns_name') and dns_name:
            current['dns_name'] = dns_name
        if not current.get('model') and model:
            current['model'] = model
        if not current.get('status') and status:
            current['status'] = status

    for cable in cables_raw:
        for prefix in ('a', 'z'):
            loc = cable.get(f'{prefix}_loc', '')
            if not _loc_in_rack(loc, room_filter, rack_filter):
                continue
            loc_statuses[loc].append(cable.get('status', ''))
            _, _, ru = _parse_loc(loc)
            _upsert_device(
                loc,
                ru,
                dns_name=cable.get(f'{prefix}_dns', ''),
                model=cable.get(f'{prefix}_model', ''),
            )

    # Supplement cable-derived devices with SITE-HOSTS records so racks with
    # sparse or one-sided cabling still show their full device list.
    for loc, host in hosts_by_loc.items():
        if not _loc_in_rack(loc, room_filter, rack_filter):
            continue
        _, _, ru = _parse_loc(loc)
        _upsert_device(
            loc,
            ru,
            dns_name=host.get('dns', ''),
            model=host.get('model', ''),
            status=host.get('status', ''),
        )

    for loc, device in devices.items():
        derived_status = _derive_device_status(loc_statuses.get(loc, []))
        if derived_status:
            device['status'] = derived_status
        elif not device.get('status'):
            device['status'] = 'Listed'

    # -- Optic summary for this rack's ports --
    # Use a full two-sided scan of ALL cables so that optics on Z-side ports
    # are captured even when the cable's A-side belongs to a different rack
    # (those cables are excluded from rack_cables to avoid duplicate cable maps).
    optic_counts = defaultdict(int)
    optic_locations = []
    for cable in cables_raw:
        a_in = _loc_in_rack(cable.get('a_loc', ''), room_filter, rack_filter)
        z_in = _loc_in_rack(cable.get('z_loc', ''), room_filter, rack_filter)
        if a_in and cable.get('a_optic'):
            optic_counts[cable['a_optic']] += 1
            optic_locations.append({
                'location': cable.get('a_loc', ''),
                'port':     cable.get('a_port', ''),
                'optic':    cable['a_optic'],
            })
        if z_in and cable.get('z_optic'):
            optic_counts[cable['z_optic']] += 1
            optic_locations.append({
                'location': cable.get('z_loc', ''),
                'port':     cable.get('z_port', ''),
                'optic':    cable['z_optic'],
            })

    # Sort by RU then port (natural sort so port2 < port10 < port11)
    optic_locations.sort(key=lambda x: (_ru_sort(_parse_loc(x['location'])[2]), _natural_key(x['port'])))

    # -- Generate label strings --
    internal_labels = [
        _cable_label(c, 'a') for c in sorted(internal,
            key=lambda c: (-_ru_sort(_parse_loc(c.get('a_loc', ''))[2]), _natural_key(c.get('a_port', ''))))
    ]
    cab_to_cab_labels = [
        _cable_label(c, 'a' if c['a_in_rack'] else 'z')
        for c in sorted(cab_to_cab, key=lambda c: (
            -_ru_sort(_parse_loc(c.get('a_loc', '') if c['a_in_rack'] else c.get('z_loc', ''))[2]),
            _natural_key(c.get('a_port', '') if c['a_in_rack'] else c.get('z_port', ''))
        ))
    ]

    # -- Sort devices by RU descending (top of rack first) --
    sorted_devices = sorted(devices.values(), key=lambda d: _ru_sort(d['ru']), reverse=True)

    # -- Serialise cable lists for JSON --
    def _serialise_cable(c):
        return {k: v for k, v in c.items() if k not in ('a_in_rack', 'z_in_rack')}

    return {
        'room':            room_input,
        'rack':            rack_input,
        'cab_type':        cab_type,
        'cab_type_summary': cab_type_summary,
        'elevation':       elevation,
        'total_cables':    len(rack_cables),
        'internal_count':  len(internal),
        'cab_to_cab_count': len(cab_to_cab),
        'devices':         sorted_devices,
        'optic_summary':   dict(sorted(optic_counts.items(), key=lambda x: x[1], reverse=True)),
        'optic_locations': optic_locations,
        'internal_cables': [_serialise_cable(c) for c in sorted(
            internal, key=lambda c: (-_ru_sort(_parse_loc(c.get('a_loc', ''))[2]), _natural_key(c.get('a_port', ''))))],
        'cab_to_cab_cables': [_serialise_cable(c) for c in sorted(
            cab_to_cab, key=lambda c: (
                -_ru_sort(_parse_loc(c.get('a_loc', '') if c['a_in_rack'] else c.get('z_loc', ''))[2]),
                _natural_key(c.get('a_port', '') if c['a_in_rack'] else c.get('z_port', ''))))],
        'internal_labels':   internal_labels,
        'cab_to_cab_labels': cab_to_cab_labels,
    }


def process_room(cutsheet_path, room_input):
    """
    Return all cables where either end is in the given room (DH),
    regardless of rack. Used for the 'Download all DH labels' export.
    """
    room_filter = room_input.strip().lower()

    wb_cut = openpyxl.load_workbook(cutsheet_path, read_only=True, data_only=True)
    cut_tab = _find_cutsheet_tab(wb_cut) or 'CUTSHEET'
    cables_raw = _read_sheet(wb_cut, cut_tab, CUTSHEET_COLS)
    wb_cut.close()

    room_cables = []
    for cable in cables_raw:
        a_in = _loc_in_room(cable.get('a_loc', ''), room_filter)
        z_in = _loc_in_room(cable.get('z_loc', ''), room_filter)
        if a_in or z_in:
            cable['a_in_room'] = a_in
            cable['z_in_room'] = z_in
            room_cables.append(cable)

    internal   = [c for c in room_cables if c['a_in_room'] and c['z_in_room']]
    cab_to_cab = [c for c in room_cables if not (c['a_in_room'] and c['z_in_room'])]

    def _serialise(c):
        return {k: v for k, v in c.items() if k not in ('a_in_room', 'z_in_room')}

    def _sort_key(c, use_a=True):
        loc = c.get('a_loc', '') if use_a else (
            c.get('a_loc', '') if c.get('a_in_room') else c.get('z_loc', ''))
        room, rack, ru = _parse_loc(loc)
        return (_natural_key(rack), _ru_sort(ru), _natural_key(c.get('a_port', '') if use_a else c.get('z_port', '')))

    return {
        'room': room_input,
        'internal_cables':   [_serialise(c) for c in sorted(internal, key=lambda c: _sort_key(c, use_a=True))],
        'cab_to_cab_cables': [_serialise(c) for c in sorted(cab_to_cab, key=lambda c: _sort_key(c, use_a=False))],
    }


# ---------------------------------------------------------------------------
# Layout workbook generator
# ---------------------------------------------------------------------------

def generate_layout_workbook(cutsheet_path, template_path, room_input):
    """
    Generate an Excel workbook with one tab per unique cab type found in
    the given room.  Each tab contains a rack layout sourced from either
    the Region Template (when provided) or the Cutsheet.
    Returns raw bytes suitable for streaming as a file download.
    """
    room_filter  = room_input.strip().lower()
    has_template = bool(template_path)
    source_label = 'Region Template' if has_template else 'Cutsheet'

    # Shared styles
    bold_font    = Font(bold=True)
    banner_fill  = PatternFill('solid', fgColor='FFF3CD')   # yellow warning
    header_fill  = PatternFill('solid', fgColor='DDEBF7')   # light blue headers
    green_fill   = PatternFill('solid', fgColor='D4EDDA')
    yellow_fill  = PatternFill('solid', fgColor='FFF3CD')

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # Load all cable data
    wb_cut = openpyxl.load_workbook(cutsheet_path, read_only=True, data_only=True)
    cut_tab = _find_cutsheet_tab(wb_cut) or 'CUTSHEET'
    cables_raw = _read_sheet(wb_cut, cut_tab, CUTSHEET_COLS)
    wb_cut.close()

    # OVERHEAD rack→cab_type map
    rack_to_cab = _overhead_rack_to_cab_type(cutsheet_path)

    # Find racks per cab type that have cable data in this room
    cab_racks = defaultdict(set)
    for cable in cables_raw:
        for loc_key in ('a_loc', 'z_loc'):
            room, rack, _ = _parse_loc(cable.get(loc_key, ''))
            if room and _room_matches(room, room_filter) and rack:
                try:
                    rack_int = int(rack.lstrip('0') or '0')
                    cab = rack_to_cab.get(rack_int)
                    if cab:
                        cab_racks[cab].add(rack_int)
                except ValueError:
                    pass

    for cab_type in sorted(cab_racks.keys()):
        racks_of_type = cab_racks[cab_type]
        ws = wb_out.create_sheet(title=cab_type[:31])
        row = 1

        # Source banner
        ws.cell(row=row, column=1, value=f'Source: {source_label}').font = bold_font
        row += 1

        if not has_template:
            c = ws.cell(row=row, column=1,
                        value='Unused standby devices will not be displayed')
            c.fill = banner_fill
            c.font = bold_font
            row += 1

        row += 1  # blank spacer

        if has_template:
            elevation = _lookup_elevation(template_path, cab_type)

            # RU values present in cable data for racks of this type
            cabled_rus = set()
            for cable in cables_raw:
                for loc_key in ('a_loc', 'z_loc'):
                    r2, rack2, ru2 = _parse_loc(cable.get(loc_key, ''))
                    if not (r2 and _room_matches(r2, room_filter) and ru2):
                        continue
                    try:
                        if int(rack2.lstrip('0') or '0') in racks_of_type:
                            cabled_rus.add(ru2)
                    except ValueError:
                        pass

            headers = ['RU', 'Device Name', 'Device Type', 'Cabling Found']
            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c_idx, value=h)
                cell.font = bold_font
                cell.fill = header_fill
            row += 1

            for item in elevation:
                ru = item['ru']
                found = bool(ru and ru in cabled_rus)
                cabling_val = 'YES' if found else ('NO' if ru else '')
                ws.cell(row=row, column=1, value=ru)
                ws.cell(row=row, column=2, value=item['device_name'])
                ws.cell(row=row, column=3, value=item['device_type'])
                found_cell = ws.cell(row=row, column=4, value=cabling_val)
                if cabling_val == 'YES':
                    found_cell.fill = green_fill
                elif cabling_val == 'NO':
                    found_cell.fill = yellow_fill
                row += 1

        else:
            # Cutsheet source: collect devices from cable data for these racks
            loc_statuses = defaultdict(list)
            seen_locs = {}
            for cable in cables_raw:
                a_room, a_rack, a_ru = _parse_loc(cable.get('a_loc', ''))
                if not (a_room and _room_matches(a_room, room_filter)):
                    continue
                try:
                    a_rack_int = int(a_rack.lstrip('0') or '0')
                except ValueError:
                    continue
                if a_rack_int not in racks_of_type:
                    continue
                loc = cable.get('a_loc', '')
                if loc:
                    loc_statuses[loc].append(cable.get('status', ''))
                if loc and a_ru and a_ru not in seen_locs:
                    seen_locs[a_ru] = {
                        'ru':    a_ru,
                        'model': cable.get('a_model', ''),
                        'loc':   loc,
                    }

            headers = ['RU', 'Model']
            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c_idx, value=h)
                cell.font = bold_font
                cell.fill = header_fill
            row += 1

            for dev in sorted(seen_locs.values(), key=lambda d: -_ru_sort(d['ru'])):
                ws.cell(row=row, column=1, value=dev['ru'])
                ws.cell(row=row, column=2, value=dev['model'])
                row += 1

        # Auto-fit columns (approximate)
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.getvalue()
